"""
Reads Lifetime Carnage Report.xlsx → docs/data/stats.json
Run after each session: python3 scripts/generate.py
"""
import pandas as pd
import json
from pathlib import Path
from collections import defaultdict

ROOT        = Path(__file__).parent.parent
EXCEL_PATH  = ROOT / 'data' / 'Lifetime Carnage Report.xlsx'
OUTPUT_PATH = ROOT / 'docs' / 'data' / 'stats.json'

TEAM_NAMES  = {0: 'Red', 1: 'Blue', 2: 'Green', 7: 'Pink', 5: 'Gold', 6: 'Brown'}
MATCHUP_NAMES = {1: 'Blue vs Red', 2: 'Pink vs Green', 3: 'Gold vs Brown'}
TEAM_COLORS = {
    'Red': '#ff4444', 'Blue': '#4488ff', 'Green': '#44dd88',
    'Pink': '#ff88cc', 'Gold': '#ffcc00', 'Brown': '#cc8855',
}
PLAYER_COLORS = {
    'BrundonDru':  '#ffcc00',
    'HDTurkie':    '#4488ff',
    'IKEMAN2684':  '#ff4444',
    'walsh696969': '#44dd88',
}
DISPLAY_NAMES = {
    'BrundonDru': 'Brundo', 'HDTurkie': 'HD',
    'IKEMAN2684': 'Ike',    'walsh696969': 'Walsh',
}
MATCHUP_PAIRS = {
    'Blue vs Red': ('Blue', 'Red'),
    'Pink vs Green': ('Pink', 'Green'),
    'Gold vs Brown': ('Gold', 'Brown'),
}

CORE_SUFFIXES = {
    'Team','Score','Kills','Deaths','Assists','Betrayals','Suicides',
    'WeaponKills','GrenadeKills','MeleeKills','OtherKills',
    'SecondsAlive','SecondsPlayed','Standing','TotalMedalCount',
    'MostKillsInARow','KilledMostPlayer','KilledMostPlayerCount',
    'MostKilledByPlayer','MostKilledByPlayerCount',
    'MostUsedWeapon','MostUsedWeaponCount',
    'OddballTime','HillTime','CamoKills','GrenadeSticks',
    'FlagScores','FlagGrabs','BombScores','BombGrabs',
    'KDR','AverageLife','DoubleKills','TripleKills','Overkills',
    'Killtaculars','KillingSprees','RunningRiots','Rampages',
    'Unfrigginbelievables','SniperSprees',
}

_TEAM_NAME_SET    = set(TEAM_NAMES.values())
_MATCHUP_NAME_SET = set(MATCHUP_NAMES.values())

# Season keys that span multiple seasons and should use lifetime labels/nite numbers
_MULTI_SEASON = (None, 'Lifetime')


# ── helpers ────────────────────────────────────────────────────────────────
def _i(v):
    try: return int(v) if v == v and v is not None else 0
    except: return 0

def _f(v):
    try: return round(float(v), 4) if v == v and v is not None else 0.0
    except: return 0.0

def _s(v):
    return str(v) if v is not None and v == v else ''

def should_combine(season_num):
    """True when suicides+betrayals are combined (pre-S6 data)."""
    if season_num in _MULTI_SEASON:
        return True
    try:
        return int(season_num) < 6
    except (TypeError, ValueError):
        return True

def _decode_team(v):
    """Accept a team-name string or a numeric code; return the name string."""
    if v is None or (v != v): return ''
    if isinstance(v, str) and v in _TEAM_NAME_SET: return v
    return TEAM_NAMES.get(_i(v), str(v))

def _decode_matchup(v):
    """Accept a matchup-name string or a numeric code; return the name string."""
    if v is None or (v != v): return ''
    if isinstance(v, str) and v in _MATCHUP_NAME_SET: return v
    code = _i(v)
    return MATCHUP_NAMES.get(code, f'Matchup {code}')

def get_players(df):
    return [c.replace('_Kills','') for c in df.columns if c.endswith('_Kills')]

def get_medal_suffixes(df, players):
    """Return ordered list of medal column suffixes (stripped of player prefix)."""
    seen, out = set(), []
    for p in players:
        prefix = p + '_'
        for col in df.columns:
            if col.startswith(prefix):
                suf = col[len(prefix):]
                if suf not in CORE_SUFFIXES and suf not in seen:
                    seen.add(suf)
                    out.append(suf)
    return out


# ── per-game construction ──────────────────────────────────────────────────
def build_games(df, players, medal_suffixes, season_num):
    games = []
    combine = should_combine(season_num)
    use_lifetime_nite = season_num in _MULTI_SEASON

    for _, row in df.iterrows():
        matchup = _decode_matchup(row.get('Matchup'))
        wt_name = _decode_team(row.get('Winning_Team'))
        lt_name = _decode_team(row.get('Losing_Team'))

        sd_raw  = row.get('Set_Decider')
        sw_raw  = row.get('Set_Winner')
        set_dec = 1 if _i(sd_raw) == 1 else 0
        set_win = _decode_team(sw_raw) if set_dec else ''

        # Collect raw per-player stats first
        raw = {}
        for p in players:
            tc = _i(row.get(f'{p}_Team'))
            raw[p] = {
                'team': TEAM_NAMES.get(tc, str(tc)),
                'team_code': tc,
                'kills': _i(row.get(f'{p}_Kills')),
                'deaths': _i(row.get(f'{p}_Deaths')),
                'assists': _i(row.get(f'{p}_Assists')),
                'betrayals': _i(row.get(f'{p}_Betrayals')),
                'suicides': _i(row.get(f'{p}_Suicides')),
                'weapon_kills': _i(row.get(f'{p}_WeaponKills')),
                'grenade_kills': _i(row.get(f'{p}_GrenadeKills')),
                'melee_kills': _i(row.get(f'{p}_MeleeKills')),
                'other_kills': _i(row.get(f'{p}_OtherKills')),
                'score': _i(row.get(f'{p}_Score')),
                'kdr': _f(row.get(f'{p}_KDR')),
                'most_kills_in_row': _i(row.get(f'{p}_MostKillsInARow')),
                'total_medals': _i(row.get(f'{p}_TotalMedalCount')),
                'avg_life': _f(row.get(f'{p}_AverageLife')),
                'seconds_played': _i(row.get(f'{p}_SecondsPlayed')),
                'medals': {suf: _i(row.get(f'{p}_{suf}'))
                           for suf in medal_suffixes
                           if _i(row.get(f'{p}_{suf}')) > 0},
            }

        # Totals for QIKE
        total_kills = sum(r['kills'] for r in raw.values()) or 1
        total_score = sum(r['score'] for r in raw.values()) or 1
        game_time   = max((raw[p]['seconds_played'] for p in players), default=0)

        # Derived per-player: KC, QIKE, spreads
        player_data = {}
        for p in players:
            r = raw[p]
            teammate_assists = sum(
                raw[q]['assists'] for q in players
                if q != p and raw[q]['team'] == r['team']
            )
            kc  = round(r['kills'] - 0.5 * teammate_assists + 0.5 * r['assists'], 2)
            qike = round(0.5 * (r['kills'] / total_kills) + 0.5 * (r['score'] / total_score), 4)
            kd_spread  = r['kills'] - r['deaths']
            kcd_spread = round(kc - r['deaths'], 2)
            sb = r['suicides'] + r['betrayals'] if combine else None

            player_data[p] = {
                **r,
                'kc': kc,
                'qike': qike,
                'kd_spread': kd_spread,
                'kcd_spread': kcd_spread,
                'suicides_betrayals': sb,
                'kills_against': {},  # filled below
            }

        # Kills-against opponents (uses KilledMostPlayer gamertag)
        for p in players:
            p_team = player_data[p]['team']
            opponents = [q for q in players if q != p and player_data[q]['team'] != p_team]
            kmp = _s(row.get(f'{p}_KilledMostPlayer'))
            kmc = _i(row.get(f'{p}_KilledMostPlayerCount'))
            kills = player_data[p]['kills']
            ka = {q: 0 for q in players if q != p}
            if kmp in opponents and kills > 0:
                ka[kmp] = kmc
                others = [o for o in opponents if o != kmp]
                if others:
                    ka[others[0]] = max(0, kills - kmc)
            elif kills > 0 and opponents:
                per, rem = divmod(kills, len(opponents))
                for i, o in enumerate(opponents):
                    ka[o] = per + (1 if i < rem else 0)
            player_data[p]['kills_against'] = ka

        # Team-level aggregates for this game
        team_map = {}
        for p in players:
            t = player_data[p]['team']
            if t not in team_map:
                team_map[t] = {'kills': 0, 'score': 0, 'assists': 0,
                               'deaths': 0, 'qike': 0.0, 'players': []}
            tm = team_map[t]
            tm['kills']   += player_data[p]['kills']
            tm['score']   += player_data[p]['score']
            tm['assists'] += player_data[p]['assists']
            tm['deaths']  += player_data[p]['deaths']
            tm['qike']    += player_data[p]['qike']
            tm['players'].append(p)

        wt_score = team_map.get(wt_name, {}).get('score', 0)
        lt_score = team_map.get(lt_name, {}).get('score', 0)

        hn_season   = _i(row.get('Halonite_Num_Season'))
        hn_lifetime = _i(row.get('Halonite_Num_Lifetime'))

        games.append({
            'game_num_season':      _i(row.get('Game_Num_Season')),
            'game_num_lifetime':    _i(row.get('Game_Num_Lifetime')),
            'halonite_num':         hn_season,
            'halonite_num_lifetime':hn_lifetime,
            'game_num_nite':        _i(row.get('Game_Num_Nite')),
            'season':               _i(row.get('Season')),
            'matchup':              matchup,
            'winning_team':         wt_name,
            'losing_team':          lt_name,
            'winning_score':        wt_score,
            'losing_score':         lt_score,
            'mov':                  wt_score - lt_score,
            'map':                  _s(row.get('Map_Manual')) or _s(row.get('Map')) or 'Unknown',
            'timestamp':            _s(row.get('Timestamp')),
            'seconds_played':       game_time,
            'set_decider':          set_dec,
            'set_winner':           set_win,
            'total_kills':          total_kills,
            'total_score':          total_score,
            'players':              player_data,
            'teams':                team_map,
        })

    return sorted(games, key=lambda g: g['game_num_lifetime'])


# ── summary ────────────────────────────────────────────────────────────────
def build_summary(games, players, season_num):
    total_seconds = sum(g['seconds_played'] for g in games)
    sets = sum(1 for g in games if g['set_decider'])

    matchup_games = defaultdict(int)
    matchup_seconds = defaultdict(int)
    for g in games:
        matchup_games[g['matchup']] += 1
        matchup_seconds[g['matchup']] += g['seconds_played']

    totals = defaultdict(int)
    for g in games:
        for p in players:
            ps = g['players'][p]
            totals['kills']        += ps['kills']
            totals['assists']      += ps['assists']
            totals['deaths']       += ps['deaths']
            totals['betrayals']    += ps['betrayals']
            totals['suicides']     += ps['suicides']
            totals['weapon_kills'] += ps['weapon_kills']
            totals['grenade_kills']+= ps['grenade_kills']
            totals['melee_kills']  += ps['melee_kills']
            totals['other_kills']  += ps['other_kills']
            totals['medals']       += ps['total_medals']
            totals['score']        += ps['score']

    total_nites = len({g['halonite_num_lifetime'] for g in games})
    return {
        'total_games':    len(games),
        'total_sets':     sets,
        'total_nites':    total_nites,
        'total_seconds':  total_seconds,
        'matchup_games':  dict(matchup_games),
        'matchup_seconds':dict(matchup_seconds),
        'combine_suicides': should_combine(season_num),
        **{k: int(v) for k, v in totals.items()},
    }


# ── player aggregates ──────────────────────────────────────────────────────
def build_aggregates(games, players, medal_suffixes, season_num):
    agg = {p: defaultdict(lambda: 0) for p in players}
    med = {p: defaultdict(int) for p in players}
    ka_agg = {p: defaultdict(int) for p in players}
    cum = {p: {k: [] for k in [
        'wl_spread','set_wl_spread','kills','deaths','assists','score',
        'kd_spread','qike','weapon_kills','grenade_kills','melee_kills','other_kills',
    ]} for p in players}
    labels = []

    run = {p: defaultdict(lambda: 0) for p in players}
    set_run = {p: defaultdict(lambda: 0) for p in players}

    use_lifetime_label = season_num in _MULTI_SEASON
    for g in games:
        labels.append(g['game_num_lifetime'] if use_lifetime_label else g['game_num_season'])
        wt = g['winning_team']
        sw = g['set_winner']

        for p in players:
            ps = g['players'][p]
            a  = agg[p]
            a['games'] += 1
            won = (ps['team'] == wt)
            a['wins']   += int(won)
            a['losses'] += int(not won)
            a['kills']  += ps['kills']
            a['deaths'] += ps['deaths']
            a['assists'] += ps['assists']
            a['betrayals'] += ps['betrayals']
            a['suicides'] += ps['suicides']
            a['weapon_kills'] += ps['weapon_kills']
            a['grenade_kills'] += ps['grenade_kills']
            a['melee_kills']  += ps['melee_kills']
            a['other_kills']  += ps['other_kills']
            a['score']  += ps['score']
            a['medals'] += ps['total_medals']
            a['kc']     += ps['kc']
            a['qike']   += ps['qike']
            a['best_kills'] = max(a['best_kills'], ps['kills'])
            a['best_kdr']   = max(a['best_kdr'],   ps['kdr'])
            a['best_score'] = max(a['best_score'],  ps['score'])
            for suf, cnt in ps['medals'].items():
                med[p][suf] += cnt
            for opp, cnt in ps.get('kills_against', {}).items():
                ka_agg[p][opp] += cnt
            if g['set_decider'] and sw:
                sw_won = (ps['team'] == sw)
                a['set_wins']   += int(sw_won)
                a['set_losses'] += int(not sw_won)
                set_run[p]['spread'] += 1 if sw_won else -1

            run[p]['wl']    += 1 if won else -1
            run[p]['kills'] += ps['kills']
            run[p]['deaths']+= ps['deaths']
            run[p]['assists']+= ps['assists']
            run[p]['score'] += ps['score']
            run[p]['kd']    += ps['kd_spread']
            run[p]['qike']  += ps['qike']
            run[p]['wk']    += ps['weapon_kills']
            run[p]['gk']    += ps['grenade_kills']
            run[p]['mk']    += ps['melee_kills']
            run[p]['ok']    += ps['other_kills']

            c = cum[p]
            c['wl_spread'].append(run[p]['wl'])
            c['set_wl_spread'].append(set_run[p]['spread'])
            c['kills'].append(run[p]['kills'])
            c['deaths'].append(run[p]['deaths'])
            c['assists'].append(run[p]['assists'])
            c['score'].append(run[p]['score'])
            c['kd_spread'].append(run[p]['kd'])
            c['qike'].append(round(run[p]['qike'], 4))
            c['weapon_kills'].append(run[p]['wk'])
            c['grenade_kills'].append(run[p]['gk'])
            c['melee_kills'].append(run[p]['mk'])
            c['other_kills'].append(run[p]['ok'])

    result = {}
    for p in players:
        a = agg[p]
        g = a['games'] or 1
        d = a['deaths'] or 1
        result[p] = {
            'games':    int(a['games']),
            'wins':     int(a['wins']),
            'losses':   int(a['losses']),
            'wl_spread':int(a['wins'] - a['losses']),
            'win_pct':  round(a['wins'] / g * 100, 1),
            'set_wins': int(a['set_wins']),
            'set_losses':int(a['set_losses']),
            'set_wl_spread': int(a['set_wins'] - a['set_losses']),
            'set_win_pct': round(a['set_wins'] / (a['set_wins']+a['set_losses'] or 1) * 100, 1),
            'total_kills':  int(a['kills']),
            'total_deaths': int(a['deaths']),
            'total_assists':int(a['assists']),
            'total_betrayals': int(a['betrayals']),
            'total_suicides':  int(a['suicides']),
            'total_weapon_kills':  int(a['weapon_kills']),
            'total_grenade_kills': int(a['grenade_kills']),
            'total_melee_kills':   int(a['melee_kills']),
            'total_other_kills':   int(a['other_kills']),
            'total_score':  int(a['score']),
            'total_medals': int(a['medals']),
            'total_kc':     round(a['kc'], 2),
            'total_qike':   round(a['qike'], 4),
            'kdr':          round(a['kills'] / d, 3),
            'avg_kills':    round(a['kills'] / g, 2),
            'avg_deaths':   round(a['deaths'] / g, 2),
            'avg_score':    round(a['score'] / g, 2),
            'avg_kc':       round(a['kc'] / g, 2),
            'avg_qike':     round(a['qike'] / g, 4),
            'kd_spread':    int(a['kills'] - a['deaths']),
            'kcd_spread':   round(a['kc'] - a['deaths'], 2),
            'best_kills':   int(a['best_kills']),
            'best_kdr':     round(a['best_kdr'], 3),
            'best_score':   int(a['best_score']),
            'combine_suicides': should_combine(season_num),
            'medals': dict(med[p]),
            'kills_against': {opp: int(ka_agg[p][opp]) for opp in players if opp != p},
            'cumulative': cum[p],
            'game_labels': labels,
        }
    return result


# ── team aggregates ────────────────────────────────────────────────────────
def build_teams(games, players, season_num=None):
    acc = defaultdict(lambda: defaultdict(lambda: 0))
    cum = {}
    run = defaultdict(lambda: defaultdict(lambda: 0))
    labels = []

    use_lifetime_label = season_num in _MULTI_SEASON
    for g in games:
        labels.append(g['game_num_lifetime'] if use_lifetime_label else g['game_num_season'])
        wt = g['winning_team']
        sw = g['set_winner']

        for p in players:
            ps = g['players'][p]
            t  = ps['team']
            a  = acc[t]
            won = (t == wt)
            a['games']   += 1
            a['wins']    += int(won)
            a['losses']  += int(not won)
            a['kills']   += ps['kills']
            a['deaths']  += ps['deaths']
            a['assists'] += ps['assists']
            a['score']   += ps['score']
            a['kc']      += ps['kc']
            a['qike']    += ps['qike']
            a['weapon_kills']  += ps['weapon_kills']
            a['grenade_kills'] += ps['grenade_kills']
            a['melee_kills']   += ps['melee_kills']
            a['other_kills']   += ps['other_kills']
            a['betrayals']     += ps['betrayals']
            a['suicides']      += ps['suicides']
            if won:
                a['total_mov'] += g['mov']
            if g['set_decider'] and sw:
                a['set_wins']   += int(t == sw)
                a['set_losses'] += int(t != sw)

            if t not in cum:
                cum[t] = {'wl_spread': [], 'set_wl_spread': [], 'qike': []}
            run[t]['wl']      += 1 if won else -1
            run[t]['qike']    += ps['qike']
            run[t]['set_wl']  += (1 if t == sw else -1) if (g['set_decider'] and sw) else 0

        for t in list(cum.keys()):
            cum[t]['wl_spread'].append(run[t]['wl'] // 2)
            cum[t]['set_wl_spread'].append(run[t]['set_wl'] // 2)
            cum[t]['qike'].append(round(run[t]['qike'], 4))

    result = {}
    for t, a in acc.items():
        g_count = a['games'] // 2 or 1
        wins    = a['wins'] // 2
        losses  = a['losses'] // 2
        deaths  = a['deaths'] or 1
        sw_ = a['set_wins'] // 2
        sl_ = a['set_losses'] // 2
        result[t] = {
            'color':       TEAM_COLORS.get(t, '#aaa'),
            'games':       g_count,
            'wins':        wins,
            'losses':      losses,
            'wl_spread':   wins - losses,
            'win_pct':     round(wins / g_count * 100, 1),
            'set_wins':    sw_,
            'set_losses':  sl_,
            'set_wl_spread': sw_ - sl_,
            'set_win_pct': round(sw_ / (sw_ + sl_ or 1) * 100, 1),
            'total_kills':  a['kills'],
            'total_deaths': a['deaths'],
            'total_assists':a['assists'],
            'total_score':  a['score'],
            'total_kc':     round(a['kc'], 2),
            'total_qike':   round(a['qike'], 4),
            'kdr':          round(a['kills'] / deaths, 3),
            'kd_spread':    a['kills'] - a['deaths'],
            'avg_kills':    round(a['kills'] / g_count, 2),
            'avg_deaths':   round(a['deaths'] / g_count, 2),
            'avg_score':    round(a['score'] / g_count, 2),
            'avg_qike':     round(a['qike'] / (a['games'] or 1), 4),
            'avg_mov':      round(a['total_mov'] / (a['wins'] or 1), 2),
            'weapon_kills':  a['weapon_kills'],
            'grenade_kills': a['grenade_kills'],
            'melee_kills':   a['melee_kills'],
            'other_kills':   a['other_kills'],
            'betrayals':     a['betrayals'],
            'suicides':      a['suicides'],
            'cumulative':    cum.get(t, {}),
            'game_labels':   labels,
        }
    return result


# ── matchups ───────────────────────────────────────────────────────────────
def build_matchups(games):
    acc = defaultdict(lambda: defaultdict(lambda: 0))
    for g in games:
        m  = g['matchup']
        ta, tb = MATCHUP_PAIRS.get(m, ('', ''))
        acc[m]['games'] += 1
        acc[m]['seconds'] += g['seconds_played']
        if g['winning_team'] == ta:
            acc[m]['a_wins'] += 1
        elif g['winning_team'] == tb:
            acc[m]['b_wins'] += 1
        if g['set_decider'] and g['set_winner']:
            sw = g['set_winner']
            if sw == ta:   acc[m]['a_set_wins'] += 1
            elif sw == tb: acc[m]['b_set_wins'] += 1

    result = {}
    for m, a in acc.items():
        ta, tb = MATCHUP_PAIRS.get(m, ('', ''))
        g = a['games'] or 1
        result[m] = {
            'team_a': ta, 'team_b': tb,
            'games':  a['games'],
            'seconds':a['seconds'],
            'a_wins': a['a_wins'], 'b_wins': a['b_wins'],
            'a_set_wins': a['a_set_wins'], 'b_set_wins': a['b_set_wins'],
            'a_win_pct': round(a['a_wins'] / g * 100, 1),
            'b_win_pct': round(a['b_wins'] / g * 100, 1),
        }
    return result


# ── maps ───────────────────────────────────────────────────────────────────
def build_maps(games, players):
    acc = defaultdict(lambda: defaultdict(lambda: 0))
    p_kills = defaultdict(lambda: defaultdict(int))
    p_score = defaultdict(lambda: defaultdict(int))
    p_games = defaultdict(lambda: defaultdict(int))
    p_wins  = defaultdict(lambda: defaultdict(int))
    team_wins  = defaultdict(lambda: defaultdict(int))
    team_games = defaultdict(lambda: defaultdict(int))
    mu_games   = defaultdict(lambda: defaultdict(int))
    mu_wins    = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for g in games:
        mn = g['map']
        if not mn or mn == 'Unknown': continue
        acc[mn]['games']   += 1
        acc[mn]['seconds'] += g['seconds_played']
        acc[mn]['mov_sum'] += g['mov']
        wt = g['winning_team']
        lt = g['losing_team']
        mu = g['matchup']
        team_wins[mn][wt]  += 1
        team_games[mn][wt] += 1
        team_games[mn][lt] += 1
        mu_games[mn][mu]   += 1
        mu_wins[mn][mu][wt]+= 1

        for p in players:
            ps = g['players'][p]
            acc[mn]['kills']        += ps['kills']
            acc[mn]['assists']      += ps['assists']
            acc[mn]['deaths']       += ps['deaths']
            acc[mn]['weapon_kills'] += ps['weapon_kills']
            acc[mn]['grenade_kills']+= ps['grenade_kills']
            acc[mn]['melee_kills']  += ps['melee_kills']
            acc[mn]['other_kills']  += ps['other_kills']
            acc[mn]['betrayals']    += ps['betrayals']
            acc[mn]['suicides']     += ps['suicides']
            acc[mn]['hill_offense'] += ps['medals'].get('Hill Offense', 0)
            acc[mn]['hill_defense'] += ps['medals'].get('Hill Defense', 0)
            p_kills[mn][p] += ps['kills']
            p_score[mn][p] += ps['score']
            p_games[mn][p] += 1
            p_wins[mn][p]  += int(ps['team'] == wt)

    result = {}
    for mn in sorted(acc.keys()):
        a  = acc[mn]
        g  = a['games'] or 1
        n  = len(players)
        tk = a['kills'] or 1
        player_avg_kills = {p: round(p_kills[mn][p] / (p_games[mn][p] or 1), 2) for p in players}
        player_avg_score = {p: round(p_score[mn][p] / (p_games[mn][p] or 1), 2) for p in players}
        player_win_pct   = {p: round(p_wins[mn][p] / (p_games[mn][p] or 1) * 100, 1) for p in players}
        tw = {t: round(team_wins[mn][t] / (team_games[mn][t] or 1) * 100, 1)
              for t in team_games[mn]}
        mu_breakdown = {}
        for mu_name, g_cnt in mu_games[mn].items():
            ta, tb = MATCHUP_PAIRS.get(mu_name, ('', ''))
            mu_breakdown[mu_name] = {
                'games':    g_cnt,
                'a_wins':   mu_wins[mn][mu_name].get(ta, 0),
                'b_wins':   mu_wins[mn][mu_name].get(tb, 0),
                'a_win_pct':round(mu_wins[mn][mu_name].get(ta,0)/(g_cnt or 1)*100,1),
                'b_win_pct':round(mu_wins[mn][mu_name].get(tb,0)/(g_cnt or 1)*100,1),
            }
        result[mn] = {
            'games':       g,
            'total_kills': a['kills'],
            'avg_mov':     round(a['mov_sum'] / g, 1),
            'avg_seconds': round(a['seconds'] / g),
            'avg_kills_per_player':  round(a['kills'] / g / n, 2),
            'avg_deaths_per_player': round(a['deaths'] / g / n, 2),
            'avg_assists_pct':       round(a['assists'] / tk * 100, 1),
            'avg_betray_suicide':    round((a['betrayals'] + a['suicides']) / g, 2),
            'weapon_pct':    round(a['weapon_kills'] / tk * 100, 1),
            'grenade_pct':   round(a['grenade_kills'] / tk * 100, 1),
            'melee_pct':     round(a['melee_kills'] / tk * 100, 1),
            'other_pct':     round(a['other_kills'] / tk * 100, 1),
            'hill_offense':     a['hill_offense'],
            'hill_defense':     a['hill_defense'],
            'hill_offense_pct': round(a['hill_offense'] / tk * 100, 1),
            'hill_defense_pct': round(a['hill_defense'] / tk * 100, 1),
            'player_avg_kills': player_avg_kills,
            'player_avg_score': player_avg_score,
            'player_win_pct':   player_win_pct,
            'team_win_pct':     tw,
            'mu_breakdown':     mu_breakdown,
            'top_player': max(player_avg_kills, key=player_avg_kills.get),
        }
    return result


# ── halo nites ─────────────────────────────────────────────────────────────
def build_halo_nites(games, players, season_num):
    # Use lifetime nite numbers for multi-season views to avoid collisions
    use_lifetime_nite = season_num in _MULTI_SEASON
    nites = defaultdict(list)
    for g in games:
        key = g['halonite_num_lifetime'] if use_lifetime_nite else g['halonite_num']
        nites[key].append(g)

    result = []
    for nite_num in sorted(nites.keys()):
        nite_games = nites[nite_num]
        p_stats = {p: defaultdict(lambda: 0) for p in players}
        game_log = []

        for g in nite_games:
            game_log.append({
                'game_num_season': g['game_num_season'],
                'game_num_nite':   g['game_num_nite'],
                'map':             g['map'],
                'matchup':         g['matchup'],
                'winning_team':    g['winning_team'],
                'seconds_played':  g['seconds_played'],
            })
            wt = g['winning_team']
            sw = g['set_winner']
            for p in players:
                ps = g['players'][p]
                pst = p_stats[p]
                pst['games']        += 1
                pst['kills']        += ps['kills']
                pst['deaths']       += ps['deaths']
                pst['assists']      += ps['assists']
                pst['score']        += ps['score']
                pst['kc']           += ps['kc']
                pst['kd_spread']    += ps['kd_spread']
                pst['kcd_spread']   += ps['kcd_spread']
                pst['qike']         += ps['qike']
                pst['weapon_kills'] += ps['weapon_kills']
                pst['grenade_kills']+= ps['grenade_kills']
                pst['melee_kills']  += ps['melee_kills']
                pst['other_kills']  += ps['other_kills']
                pst['medals']       += ps['total_medals']
                pst['betrayals']    += ps['betrayals']
                pst['suicides']     += ps['suicides']
                pst['wins']         += int(ps['team'] == wt)
                pst['losses']       += int(ps['team'] != wt)
                if g['set_decider'] and sw:
                    pst['set_wins']  += int(ps['team'] == sw)
                    pst['set_losses']+= int(ps['team'] != sw)

        player_summary = {}
        for p in players:
            st = p_stats[p]
            g  = st['games'] or 1
            player_summary[p] = {
                'games':        int(st['games']),
                'wins':         int(st['wins']),
                'losses':       int(st['losses']),
                'wl_spread':    int(st['wins'] - st['losses']),
                'win_pct':      round(st['wins'] / g * 100, 1),
                'set_wins':     int(st['set_wins']),
                'set_losses':   int(st['set_losses']),
                'score':        int(st['score']),
                'kills':        int(st['kills']),
                'deaths':       int(st['deaths']),
                'assists':      int(st['assists']),
                'kd_spread':    int(st['kd_spread']),
                'kc':           round(st['kc'], 2),
                'kcd_spread':   round(st['kcd_spread'], 2),
                'avg_qike':     round(st['qike'] / g, 4),
                'weapon_kills': int(st['weapon_kills']),
                'grenade_kills':int(st['grenade_kills']),
                'melee_kills':  int(st['melee_kills']),
                'other_kills':  int(st['other_kills']),
                'medals':       int(st['medals']),
                'betrayals':    int(st['betrayals']),
                'suicides':     int(st['suicides']),
                'combine_suicides': should_combine(season_num),
            }

        result.append({
            'nite_num':     nite_num,
            'game_count':   len(nite_games),
            'game_log':     game_log,
            'player_stats': player_summary,
        })
    return result


# ── records ────────────────────────────────────────────────────────────────
def build_records(games, players, season_num):
    combine = should_combine(season_num)

    def make_tracker(minimize=False):
        return {'val': float('inf') if minimize else float('-inf'), 'instances': []}

    def check(tracker, val, info, minimize=False):
        if val is None: return
        better = (val < tracker['val']) if minimize else (val > tracker['val'])
        if better:
            tracker['val'] = val
            tracker['instances'] = [info]
        elif val == tracker['val']:
            tracker['instances'].append(info)

    pp = {k: make_tracker() for k in [
        'most_points','most_kills','most_assists','fewest_deaths',
        'most_weapon_kills','most_grenade_kills','most_melee_kills','most_other_kills',
        'greatest_spread','longest_spree',
        'longest_win_streak','longest_set_win_streak',
    ]}
    pp['fewest_deaths'] = make_tracker(minimize=True)

    pn = {k: make_tracker(True if k != 'most_deaths' and k != 'most_betrayals' and k != 'most_suicides' and k != 'most_betray_suicide' else False) for k in [
        'fewest_points','fewest_kills','most_deaths','lowest_spread',
        'most_betrayals','most_suicides','most_betray_suicide',
    ]}
    pn['fewest_points']  = make_tracker(minimize=True)
    pn['fewest_kills']   = make_tracker(minimize=True)
    pn['lowest_spread']  = make_tracker(minimize=True)
    pn['most_deaths']    = make_tracker(minimize=False)
    pn['most_betrayals'] = make_tracker(minimize=False)
    pn['most_suicides']  = make_tracker(minimize=False)
    pn['most_betray_suicide']   = make_tracker(minimize=False)
    pn['longest_loss_streak']    = make_tracker(minimize=False)
    pn['longest_set_loss_streak']= make_tracker(minimize=False)

    streak_gw = {p: 0 for p in players}
    streak_gl = {p: 0 for p in players}
    streak_sw = {p: 0 for p in players}
    streak_sl = {p: 0 for p in players}

    tp = {k: make_tracker() for k in ['most_kills','most_assists','fewest_deaths','greatest_spread']}
    tp['fewest_deaths'] = make_tracker(minimize=True)

    tn = {k: make_tracker(minimize=(k in ['fewest_points','fewest_kills','lowest_spread'])) for k in [
        'fewest_points','fewest_kills','most_deaths','lowest_spread',
    ]}
    tn['most_deaths'] = make_tracker(minimize=False)

    for g in games:
        info_base = {
            'game_num_season': g['game_num_season'],
            'halonite_num':    g['halonite_num'],
            'season':          g['season'],
            'map':             g['map'],
            'matchup':         g['matchup'],
        }

        for p in players:
            ps = g['players'][p]
            info = {**info_base, 'player': p, 'display': DISPLAY_NAMES.get(p, p)}
            sb = ps['betrayals'] + ps['suicides']

            check(pp['most_points'],       ps['score'],             info)
            check(pp['most_kills'],        ps['kills'],             info)
            check(pp['most_assists'],      ps['assists'],           info)
            check(pp['fewest_deaths'],     ps['deaths'],            info, minimize=True)
            check(pp['most_weapon_kills'], ps['weapon_kills'],      info)
            check(pp['most_grenade_kills'],ps['grenade_kills'],     info)
            check(pp['most_melee_kills'],  ps['melee_kills'],       info)
            check(pp['most_other_kills'],  ps['other_kills'],       info)
            check(pp['greatest_spread'],   ps['kd_spread'],         info)
            check(pp['longest_spree'],     ps['most_kills_in_row'], info)

            check(pn['fewest_points'],  ps['score'],      info, minimize=True)
            check(pn['fewest_kills'],   ps['kills'],      info, minimize=True)
            check(pn['most_deaths'],    ps['deaths'],     info)
            check(pn['lowest_spread'],  ps['kd_spread'],  info, minimize=True)
            if combine:
                check(pn['most_betray_suicide'], sb, info)
            else:
                check(pn['most_betrayals'], ps['betrayals'], info)
                check(pn['most_suicides'],  ps['suicides'],  info)

            won = ps['team'] == g['winning_team']
            if won:
                streak_gw[p] += 1; streak_gl[p] = 0
            else:
                streak_gl[p] += 1; streak_gw[p] = 0
            check(pp['longest_win_streak'],  streak_gw[p], info)
            check(pn['longest_loss_streak'], streak_gl[p], info)
            sw_team = g.get('set_winner', '')
            if g.get('set_decider') and sw_team:
                if ps['team'] == sw_team:
                    streak_sw[p] += 1; streak_sl[p] = 0
                else:
                    streak_sl[p] += 1; streak_sw[p] = 0
                check(pp['longest_set_win_streak'],  streak_sw[p], info)
                check(pn['longest_set_loss_streak'], streak_sl[p], info)

        for team_name, tm in g['teams'].items():
            spread = tm['kills'] - tm['deaths']
            tinfo = {**info_base, 'team': team_name}
            check(tp['most_kills'],      tm['kills'],   tinfo)
            check(tp['most_assists'],    tm['assists'],  tinfo)
            check(tp['fewest_deaths'],   tm['deaths'],  tinfo, minimize=True)
            check(tp['greatest_spread'], spread,        tinfo)
            check(tn['fewest_points'],   tm['score'],   tinfo, minimize=True)
            check(tn['fewest_kills'],    tm['kills'],   tinfo, minimize=True)
            check(tn['most_deaths'],     tm['deaths'],  tinfo)
            check(tn['lowest_spread'],   spread,        tinfo, minimize=True)

    def finalize(tracker):
        return {'value': tracker['val'] if tracker['val'] not in (float('inf'), float('-inf')) else None,
                'instances': tracker['instances']}

    return {
        'player_positive': {k: finalize(v) for k, v in pp.items()},
        'player_negative': {k: finalize(v) for k, v in pn.items()},
        'team_positive':   {k: finalize(v) for k, v in tp.items()},
        'team_negative':   {k: finalize(v) for k, v in tn.items()},
        'combine_suicides': combine,
    }


# ── personal bests ──────────────────────────────────────────────────────────
def build_personal_bests(games, players):
    categories = {
        'most_kills': ('kills', False),
        'most_points': ('score', False),
        'most_assists': ('assists', False),
        'best_kdr': ('kdr', False),
        'best_spread': ('kd_spread', False),
        'fewest_deaths': ('deaths', True),
    }
    streak_cats = ['longest_win_streak','longest_loss_streak','longest_set_win_streak','longest_set_loss_streak']
    all_keys = list(categories.keys()) + streak_cats
    bests = {p: {k: {'val': None, 'instances': []} for k in all_keys} for p in players}

    sgw = {p: 0 for p in players}
    sgl = {p: 0 for p in players}
    ssw = {p: 0 for p in players}
    ssl = {p: 0 for p in players}

    for g in games:
        for p in players:
            ps = g['players'][p]
            info = {
                'player': p, 'display': DISPLAY_NAMES.get(p, p),
                'game_num_season': g['game_num_season'],
                'halonite_num': g['halonite_num'],
                'season': g['season'],
                'map': g['map'],
                'matchup': g['matchup'],
            }
            for cat, (field, minimize) in categories.items():
                val = ps.get(field)
                if val is None: continue
                best = bests[p][cat]
                if best['val'] is None:
                    best['val'] = val
                    best['instances'] = [{**info, 'value': val}]
                elif (val < best['val']) if minimize else (val > best['val']):
                    best['val'] = val
                    best['instances'] = [{**info, 'value': val}]
                elif val == best['val']:
                    best['instances'].append({**info, 'value': val})

            won = ps['team'] == g['winning_team']
            if won: sgw[p] += 1; sgl[p] = 0
            else:   sgl[p] += 1; sgw[p] = 0
            if g.get('set_decider') and g.get('set_winner'):
                if ps['team'] == g['set_winner']: ssw[p] += 1; ssl[p] = 0
                else:                              ssl[p] += 1; ssw[p] = 0

            for cat, streak_val in [
                ('longest_win_streak',      sgw[p]),
                ('longest_loss_streak',     sgl[p]),
                ('longest_set_win_streak',  ssw[p]),
                ('longest_set_loss_streak', ssl[p]),
            ]:
                if streak_val == 0: continue
                best = bests[p][cat]
                entry = {**info, 'value': streak_val}
                if best['val'] is None or streak_val > best['val']:
                    best['val'] = streak_val; best['instances'] = [entry]
                elif streak_val == best['val']:
                    best['instances'].append(entry)

    return bests


# ── main ───────────────────────────────────────────────────────────────────
def build_season_data(games, players, medal_suffixes, season_num):
    return {
        'season_num':  season_num,
        'combine_suicides': should_combine(season_num),
        'summary':     build_summary(games, players, season_num),
        'games':       games,
        'aggregates':  build_aggregates(games, players, medal_suffixes, season_num),
        'teams':       build_teams(games, players, season_num),
        'matchups':    build_matchups(games),
        'maps':        build_maps(games, players),
        'halo_nites':  build_halo_nites(games, players, season_num),
        'records':     build_records(games, players, season_num),
        'personal_bests': build_personal_bests(games, players),
    }


def main():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    df = pd.read_excel(EXCEL_PATH, engine='openpyxl')
    df = df[df['Game_Num_Season'].notna()].reset_index(drop=True)

    # Patch formula-based columns from data_only workbook (covers S6 ArrayFormula cells)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    PATCH_COLS = [
        'Matchup', 'Winning_Team', 'Losing_Team', 'Set_Decider', 'Set_Winner',
        'Halonite_Num_Lifetime', 'Game_Num_Lifetime', 'Game_Num_Season',
        'Game_Num_Nite', 'Game_Num_Set',
    ]
    patch_idxs = {col: headers.index(col) for col in PATCH_COLS if col in headers}
    gns_idx = headers.index('Game_Num_Season')
    game_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                 if r[gns_idx] is not None]
    for i, gr in enumerate(game_rows):
        if i >= len(df): break
        for col, idx in patch_idxs.items():
            df.at[i, col] = gr[idx]

    players        = get_players(df)
    medal_suffixes = get_medal_suffixes(df, players)

    print(f'Players:  {players}')
    print(f'Games:    {len(df)}')
    print(f'Medals:   {len(medal_suffixes)} columns per player')

    all_games = build_games(df, players, medal_suffixes, season_num=None)

    # Individual seasons
    seasons = sorted(df['Season'].dropna().unique())
    season_data = {}
    for s in seasons:
        sn = int(s)
        s_df = df[df['Season'] == s].reset_index(drop=True)
        s_games = build_games(s_df, players, medal_suffixes, season_num=sn)
        season_data[f'S{sn}'] = build_season_data(s_games, players, medal_suffixes, sn)
        print(f'  S{sn}: {len(s_games)} games')

    # Lifetime (all seasons combined)
    season_data['Lifetime'] = build_season_data(all_games, players, medal_suffixes, season_num='Lifetime')

    season_keys = ['Lifetime'] + [f'S{int(s)}' for s in seasons]

    out = {
        'seasons':       season_keys,
        'players':       players,
        'display_names': DISPLAY_NAMES,
        'player_colors': PLAYER_COLORS,
        'team_colors':   TEAM_COLORS,
        'medal_columns': medal_suffixes,
        'data':          season_data,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(out, f, separators=(',', ':'), default=str)
    size = OUTPUT_PATH.stat().st_size / 1024
    print(f'Output:   {OUTPUT_PATH}  ({size:.0f} KB)')


if __name__ == '__main__':
    main()
